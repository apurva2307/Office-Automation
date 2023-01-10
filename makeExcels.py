from openpyxl import load_workbook, Workbook
from data import get_owe_data
from openpyxl.styles import Alignment
import logging, sys, os, shutil, time, datetime

logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter("%(asctime)s:: %(levelname)s: %(message)s")
file_handler = logging.FileHandler("files/office.log")
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)
stdout_handler = logging.StreamHandler(sys.stdout)
logger.addHandler(stdout_handler)


def make_excel(month, puList, isCrore=True):
    divider = 10000 if isCrore else 1
    res = get_owe_data(month)
    logger.info(res["data1"]["PU1"]["budget"])

    if not res:
        print("No data is available for given input.")
        return
    if "msg" in res.keys():
        print(res["msg"])
        return
    data = res["data1"]
    # print(data)
    wb = load_workbook("customFilePU.xlsx")
    if month[:3] in ["JAN", "FEB", "MAR"]:
        lastYear = int(month[3:]) - 1
    else:
        lastYear = int(month[3:])
    customSheet = wb["Sheet1"]
    lastYearData = get_owe_data(f"MAR{lastYear}")["data1"]
    customSheet.cell(1, 10).value = "Fig in crore" if isCrore else "Fig in thousand"
    customSheet.cell(3, 2).value = f"20{lastYear-1}-{lastYear}"
    customSheet.cell(3, 3).value = f"20{lastYear}-{lastYear+1}"
    customSheet.cell(3, 4).value = f"{month[:3]}' {int(month[3:])-1}"
    customSheet.cell(3, 5).value = f"{month[:3]}' {month[3:]}"
    customSheet.cell(3, 6).value = f"{month[:3]}' {month[3:]}"
    for index, pu in enumerate(puList):
        budget = data[pu]["budget"]
        toEndActualsCoppy = data[pu]["toEndActualsCoppy"]
        toEndBp = data[pu]["toEndBp"]
        toEndActuals = data[pu]["toEndActuals"]
        lastFullYearActuals = lastYearData[pu]["toEndActuals"]
        row = index + 4
        customSheet.cell(row, 1).value = pu
        customSheet.cell(row, 2).value = round(lastFullYearActuals[-1] / divider, 2)
        customSheet.cell(row, 3).value = round(budget[-1] / divider, 2)
        customSheet.cell(row, 4).value = round(toEndActualsCoppy[-1] / divider, 2)
        customSheet.cell(row, 5).value = round(toEndBp[-1] / divider, 2)
        customSheet.cell(row, 6).value = round(toEndActuals[-1] / divider, 2)
        customSheet.cell(row, 7).value = f"=F{row}-E{row}"
        customSheet.cell(row, 8).value = f"=G{row}/E{row}"
        customSheet.cell(row, 9).value = f"=F{row}-D{row}"
        customSheet.cell(row, 10).value = f"=I{row}/D{row}"
        customSheet.cell(row, 11).value = f"=F{row}/C{row}"

    wb.save(f"PUwisedetails.xlsx")


def make_excel_month_wise(month, puList, isCrore=True):
    divider = 10000 if isCrore else 1
    months = [
        "APR",
        "MAY",
        "JUN",
        "JUL",
        "AUG",
        "SEP",
        "OCT",
        "NOV",
        "DEC",
        "JAN",
        "FEB",
        "MAR",
    ]
    numMonths = months.index(month[:3])
    wb = Workbook()
    ws = wb.active
    if month[:3] in ["JAN", "FEB", "MAR"]:
        lastYear = int(month[3:]) - 1
    else:
        lastYear = int(month[3:])
    lastYearData = get_owe_data(f"MAR{lastYear}")["data1"]
    ws.cell(1, 10).value = "Fig in crore" if isCrore else "Fig in Thousand"
    ws.cell(2, 3).value = f"20{lastYear-1}-{lastYear}"
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=numMonths + 3 + 1)
    ws.cell(2, 3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(2, 2).value = "Actuals"
    ws.cell(2, numMonths + 4 + 1).value = f"20{lastYear}-{lastYear+1}"
    ws.merge_cells(
        start_row=2,
        start_column=numMonths + 4 + 1,
        end_row=2,
        end_column=numMonths * 2 + 4 + 2,
    )
    ws.cell(2, numMonths + 4 + 1).alignment = Alignment(
        horizontal="center", vertical="center"
    )
    ws.cell(3, 2).value = f"20{lastYear-1}-{lastYear}"
    i = 0
    while i <= numMonths:
        ws.cell(3, 3 + i).value = months[i]
        ws.cell(3, numMonths + 5 + i).value = months[i]
        if i == 0:
            data = get_owe_data(f"{months[i]}{lastYear}")["data1"]
            global prevMonthdata
            prevMonthdata = data
            for index, pu in enumerate(puList):
                # budget = data[pu]["budget"]
                toEndActualsCoppy = data[pu]["toEndActualsCoppy"]
                # toEndBp = data[pu]["toEndBp"]
                toEndActuals = data[pu]["toEndActuals"]
                lastFullYearActuals = lastYearData[pu]["toEndActuals"]

                row = index + 4
                ws.cell(row, 2).value = round(lastFullYearActuals[-1] / divider, 2)
                ws.cell(row, 3).value = round(toEndActualsCoppy[-1] / divider, 2)
                ws.cell(row, numMonths + 5).value = round(toEndActuals[-1] / divider, 2)
                # pu headings
                ws.cell(row, 1).value = pu

        else:
            if months[i] in ["JAN", "FEB", "MAR"]:
                lastYear = month[3:]
            data = get_owe_data(f"{months[i]}{lastYear}")["data1"]
            for index, pu in enumerate(puList):
                # prevMonthdata = get_owe_data(f"{months[i-1]}{lastYear}")["data1"]
                # budget = data[pu]["budget"]
                toEndActualsCoppy = data[pu]["toEndActualsCoppy"]
                prevtoEndActualsCoppy = prevMonthdata[pu]["toEndActualsCoppy"]
                # toEndBp = data[pu]["toEndBp"]
                toEndActuals = data[pu]["toEndActuals"]
                prevtoEndActuals = prevMonthdata[pu]["toEndActuals"]
                row = index + 4
                ws.cell(row, 3 + i).value = round(
                    toEndActualsCoppy[-1] / divider, 2
                ) - round(prevtoEndActualsCoppy[-1] / divider, 2)
                ws.cell(row, numMonths + 5 + i).value = round(
                    toEndActuals[-1] / divider, 2
                ) - round(prevtoEndActuals[-1] / divider, 2)
            prevMonthdata = data
        if i == numMonths:
            ws.cell(3, 3 + i + 1).value = "Total"
            ws.cell(3, numMonths + 5 + i + 1).value = "Total"
            for index, pu in enumerate(puList):
                prevtoEndActualsCoppy = prevMonthdata[pu]["toEndActualsCoppy"]
                prevtoEndActuals = prevMonthdata[pu]["toEndActuals"]
                row = index + 4
                ws.cell(row, 3 + i + 1).value = round(
                    prevtoEndActualsCoppy[-1] / divider, 2
                )
                ws.cell(row, numMonths + 5 + i + 1).value = round(
                    prevtoEndActuals[-1] / divider, 2
                )

        i += 1

    wb.save(f"monthWiseDetails.xlsx")


if __name__ == "__main__":
    # make_excel("OCT22", ["PU1", "PU10", "PU11", "PU15", "PU32"], isCrore=True)
    # os.replace("abc.txt", "files/abc.txt")

    # shutil.move("files/abc.txt", "abc.txt")
    # make_excel_month_wise(
    #     "OCT22",
    #     ["PU1", "PU10", "PU11", "PU15", "PU31", "STAFF", "PU32", "PU27", "PU28"],
    #     isCrore=True,
    # )
    print(datetime.datetime.today().date())
    print(type(f"{datetime.datetime.today().date()}"))
    pass
