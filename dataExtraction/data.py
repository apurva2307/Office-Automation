from openpyxl import load_workbook
from puList import getPUList
from helpers import sanitizeValues, sanitizePercentValues
from decouple import config
import requests, json


def extractData(filePath):
    wb = load_workbook(filePath, data_only=True)
    detailedPuSheet = wb["PU Wise OWE"]
    # print(detailedPuSheet["B5"].value)
    # print(detailedPuSheet.cell(row=14, column=2).value)
    puList = getPUList()
    total = {}
    for column in range(3, len(puList) + 3):
        budget = []
        toEndBp = []
        toEndActualsCoppy = []
        toEndActuals = []
        varAcBp = []
        varAcBpPercent = []
        varAcCoppy = []
        varAcCoppyPercent = []
        budgetUtilization = []
        remainingBudget = []
        for i in range(5, 127, 11):
            budget = [*budget, detailedPuSheet.cell(row=i, column=column).value]
            toEndBp = [*toEndBp, detailedPuSheet.cell(row=i + 1, column=column).value]
            toEndActualsCoppy = [
                *toEndActualsCoppy,
                detailedPuSheet.cell(row=i + 2, column=column).value,
            ]
            toEndActuals = [
                *toEndActuals,
                detailedPuSheet.cell(row=i + 3, column=column).value,
            ]
            varAcBp = [*varAcBp, detailedPuSheet.cell(row=i + 4, column=column).value]
            varAcBpPercent = [
                *varAcBpPercent,
                detailedPuSheet.cell(row=i + 5, column=column).value,
            ]
            varAcCoppy = [
                *varAcCoppy,
                detailedPuSheet.cell(row=i + 6, column=column).value,
            ]
            varAcCoppyPercent = [
                *varAcCoppyPercent,
                detailedPuSheet.cell(row=i + 7, column=column).value,
            ]
            budgetUtilization = [
                *budgetUtilization,
                detailedPuSheet.cell(row=i + 8, column=column).value,
            ]
            remainingBudget = [
                *remainingBudget,
                detailedPuSheet.cell(row=i + 9, column=column).value,
            ]

        total[puList[column - 3]] = {
            "budget": sanitizeValues(budget),
            "toEndBp": sanitizeValues(toEndBp),
            "toEndActualsCoppy": sanitizeValues(toEndActualsCoppy),
            "toEndActuals": sanitizeValues(toEndActuals),
            "varAcBp": sanitizeValues(varAcBp),
            "varAcBpPercent": sanitizePercentValues(varAcBpPercent),
            "varAcCoppy": sanitizeValues(varAcCoppy),
            "varAcCoppyPercent": sanitizePercentValues(varAcCoppyPercent),
            "budgetUtilization": sanitizePercentValues(budgetUtilization),
            "remainingBudget": sanitizeValues(remainingBudget),
        }
    return total


def addToDatabase(month):
    registerURL = "https://e-commerce-api-apurva.herokuapp.com/api/v1/telebot/NCRAccountsBot/postData"
    data1 = extractData(f"../files/OWE-{month.upper()}.xlsx")
    payload = {
        "month": f"{month.upper()}",
        "type": "OWE",
        "data1": data1,
    }
    resp = requests.post(registerURL, json=payload)
    return resp.json()


def updateToDatabase(month):
    registerURL = "https://e-commerce-api-apurva.herokuapp.com/api/v1/telebot/NCRAccountsBot/updateData"
    data1 = extractData(f"../files/OWE-{month.upper()}.xlsx")
    headers = {"token": "Zr4u7x!A%C*F-JaNdRgUkXp2s5v8y/B?"}
    payload = {
        "month": f"{month.upper()}",
        "type": "OWE",
        "data1": data1,
    }
    resp = requests.post(registerURL, json=payload, headers=headers)
    return resp.json()


if __name__ == "__main__":
    months = ["APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
    for month in months:
        updateToDatabase(f"{month}21")
    print("done")
