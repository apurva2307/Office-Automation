from openpyxl import load_workbook
from data import extractData

FRMonth = "JAN22"
filePath = f"files/OWE-{FRMonth}.xlsx"
monthdata = extractData(filePath)


def make_excel(month, pu, data):
    wb = load_workbook("customFile.xlsx")
    customSheet = wb["Sheet1"]
    customSheet.cell(1, 4).value = f"{pu}"
    customSheet.cell(1, 10).value = "Fig in crore"
    customSheet.cell(3, 4).value = f"{month[:3]}' {int(month[3:])-1}"
    customSheet.cell(3, 5).value = f"{month[:3]}' {month[3:]}"
    customSheet.cell(3, 6).value = f"{month[:3]}' {month[3:]}"
    budget = data[pu]["budget"]
    toEndActualsCoppy = data[pu]["toEndActualsCoppy"]
    toEndBp = data[pu]["toEndBp"]
    toEndActuals = data[pu]["toEndActuals"]

    for val in range(4, 16):
        customSheet.cell(val, 3).value = round(budget[val - 4] / 10000, 2)
    for val in range(4, 16):
        customSheet.cell(val, 4).value = round(toEndActualsCoppy[val - 4] / 10000, 2)
    for val in range(4, 16):
        customSheet.cell(val, 5).value = round(toEndBp[val - 4] / 10000, 2)
    for val in range(4, 16):
        customSheet.cell(val, 6).value = round(toEndActuals[val - 4] / 10000, 2)
    wb.save(f"{pu}.xlsx")


if __name__ == "__main__":
    make_excel("JAN22", "PU32", monthdata)
