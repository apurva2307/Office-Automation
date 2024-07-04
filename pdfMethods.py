import tabula
from openpyxl import load_workbook, Workbook


def make_owe_ir(filepath, month):
    tables = tabula.read_pdf(filepath, pages=[1, 2, 3])
    print(type(tables))
    wb = load_workbook("files\OWE-Review-IR-Master.xlsx")
    sheet = wb["Sheet1"]
    mon = month.split(" ")[0]
    year = month.split(" ")[1]
    sheet.cell(3, 4).value = f"Actuals upto {mon} {int(year) - 1}"
    sheet.cell(3, 5).value = f"Actuals upto {mon} {year}"
    j = 4
    for table in tables:
        print(type(table))
        table.columns = [
            "Particulars",
            "D-03",
            "D-04",
            "D-05",
            "D-06",
            "D-07",
            "D-08",
            "D-09",
            "D-10",
            "D-11",
            "D-12",
            "D-12N",
            "D-13",
            "Total",
        ]
        table = table.iloc[2:]
        table = table[["Particulars", "Total"]]
        for index in range(len(table)):
            if table.Particulars.iloc[index] in ["TN - TOTAL", "NET"]:
                continue
            if table.Particulars.iloc[index] == "FMG":
                sheet.cell(j, 3).value = int(table.Total.iloc[index])
                sheet.cell(j, 3).number_format = "#,##0"
            if table.Particulars.iloc[index] == "ACT":
                sheet.cell(j, 5).value = int(table.Total.iloc[index])
                sheet.cell(j, 5).number_format = "#,##0"
            if table.Particulars.iloc[index] == "COPPY":
                sheet.cell(j, 2).value = "IR"
                sheet.cell(j, 4).value = int(table.Total.iloc[index])
                sheet.cell(j, 4).number_format = "#,##0"
            if len(table.Particulars.iloc[index]) >= 13:
                rly = table.Particulars.iloc[index].split(" ")[2]
                sheet.cell(j, 2).value = rly
                sheet.cell(j, 4).value = int(table.Total.iloc[index])
                sheet.cell(j, 4).number_format = "#,##0"
                j += 1

    wb.save(f"OWE_Review_IR.xlsx")


if __name__ == "__main__":
    make_owe_ir("files\RepFinRevController.pdf", "March 2023")
# tabula.convert_into(
#     "files\RepFinRevControl.pdf", "output.csv", output_format="csv", pages="all"
# )
