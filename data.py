from openpyxl import load_workbook
from dataExtraction.puList import getPUList, getPHs, getPHsMap
from dataExtraction.helpers import *
import requests, json, jwt, datetime
from decouple import config
import pandas as pd
from dateutil import tz


token = config("TOKEN")
ncr_data_url = config("NCR_DATA_URL")


def gen_token(token):
    encodedToken = jwt.encode(
        {
            "name": "shailendra",
            "exp": datetime.datetime.now(tz=tz.gettz("Asia/Kolkata"))
            + datetime.timedelta(seconds=300),
        },
        token,
        algorithm="HS256",
    )
    return encodedToken


def extractData(filePath):
    wb = load_workbook(filePath, data_only=True)
    detailedPuSheet = wb["PU Wise OWE"]
    # print(detailedPuSheet["B5"].value)
    # print(detailedPuSheet.cell(row=14, column=2).value)
    puList = getPUList()
    total = {}
    for column in range(3, len(puList) + 3):
        budget = []
        toEndBp = []
        toEndActualsCoppy = []
        toEndActuals = []
        varAcBp = []
        varAcBpPercent = []
        varAcCoppy = []
        varAcCoppyPercent = []
        budgetUtilization = []
        remainingBudget = []
        for i in range(5, 127, 11):
            budget = [*budget, detailedPuSheet.cell(row=i, column=column).value]
            toEndBp = [*toEndBp, detailedPuSheet.cell(row=i + 1, column=column).value]
            toEndActualsCoppy = [
                *toEndActualsCoppy,
                detailedPuSheet.cell(i + 2, column).value,
            ]
            toEndActuals = [
                *toEndActuals,
                detailedPuSheet.cell(row=i + 3, column=column).value,
            ]
            varAcBp = [*varAcBp, detailedPuSheet.cell(row=i + 4, column=column).value]
            varAcBpPercent = [
                *varAcBpPercent,
                detailedPuSheet.cell(row=i + 5, column=column).value,
            ]
            varAcCoppy = [
                *varAcCoppy,
                detailedPuSheet.cell(row=i + 6, column=column).value,
            ]
            varAcCoppyPercent = [
                *varAcCoppyPercent,
                detailedPuSheet.cell(row=i + 7, column=column).value,
            ]
            budgetUtilization = [
                *budgetUtilization,
                detailedPuSheet.cell(row=i + 8, column=column).value,
            ]
            remainingBudget = [
                *remainingBudget,
                detailedPuSheet.cell(row=i + 9, column=column).value,
            ]

        total[puList[column - 3]] = {
            "budget": sanitizeValues(budget),
            "toEndBp": sanitizeValues(toEndBp),
            "toEndActualsCoppy": sanitizeValues(toEndActualsCoppy),
            "toEndActuals": sanitizeValues(toEndActuals),
            "varAcBp": sanitizeValues(varAcBp),
            "varAcBpPercent": sanitizePercentValues(varAcBpPercent),
            "varAcCoppy": sanitizeValues(varAcCoppy),
            "varAcCoppyPercent": sanitizePercentValues(varAcCoppyPercent),
            "budgetUtilization": sanitizePercentValues(budgetUtilization),
            "remainingBudget": sanitizeValues(remainingBudget),
        }
    return total


def extractDataSummary(filePath):
    wb = load_workbook(filePath, data_only=True)
    detailedPuSheet = wb["Sheet1"]
    result = {}
    columns = [3, 6, 8, 9, 10, 11, 12, 13, 14, 15]
    columns1 = [3, 6, 8, 9, 11, 12, 13]
    rows = [5, 6, 42, 49, 60, 62, 63, 67, 109, 110, 111, 115, 116, 117]
    rowsMap = [
        "Staff",
        "Non-Staff",
        "D-Traction",
        "E-Traction",
        "E-Office",
        "HSD-Civil",
        "HSD-Gen",
        "Lease",
        "IRCA",
        "IRFA",
        "IRFC",
        "Coach-C",
        "Station-C",
        "Colony-C",
    ]
    for index, row in enumerate(rows):
        data = []
        if row < 88:
            for column in columns:
                if column == 12 or column == 14 or column == 15:
                    data = [*data, sntzSigVPer(detailedPuSheet.cell(row, column).value)]
                else:
                    data = [*data, sntzSigV(detailedPuSheet.cell(row, column).value)]
        else:
            for column in columns1:
                if column == 12 or column == 13:
                    data = [
                        *data,
                        sntzSigVPer(detailedPuSheet.cell(row, column).value),
                    ]
                else:
                    data = [*data, sntzSigV(detailedPuSheet.cell(row, column).value)]
        result[f"{rowsMap[index].upper()}"] = data

    return result


def extractDataCapex(filePath, sheet):
    wb = load_workbook(filePath, data_only=True)
    detailedPuSheet = wb[sheet]
    result = {}
    columns = [3, 4, 5, 7, 8, 9, 11, 12, 13]
    phs = getPHs()
    phsMap = getPHsMap()

    def phData(ph, rowRange, rowMap):
        for index, row in enumerate(rowRange):
            con = []
            open = []
            ncr = []
            for column in columns:
                if column < 6:
                    if column == 5:
                        con = [
                            *con,
                            sntzSigVPer(detailedPuSheet.cell(row, column).value),
                        ]
                        if ph == "EBR-P":
                            break
                    else:
                        con = [*con, sntzSigV(detailedPuSheet.cell(row, column).value)]
                if column < 10 and column > 6:
                    if column == 9:
                        open = [
                            *open,
                            sntzSigVPer(detailedPuSheet.cell(row, column).value),
                        ]
                    else:
                        open = [
                            *open,
                            sntzSigV(detailedPuSheet.cell(row, column).value),
                        ]
                if column < 14 and column > 10:
                    if column == 13:
                        ncr = [
                            *ncr,
                            sntzSigVPer(detailedPuSheet.cell(row, column).value),
                        ]
                    else:
                        ncr = [*ncr, sntzSigV(detailedPuSheet.cell(row, column).value)]
            if index == 0:
                if ph == "EBR-P":
                    result[f"{ph}"] = {f"{rowMap[index]}": {"NCR": con}}
                else:
                    result[f"{ph}"] = {
                        f"{rowMap[index]}": {"OPEN": open, "CON": con, "NCR": ncr}
                    }
            else:
                if ph == "EBR-P":
                    result[ph] = {
                        **result[ph],
                        f"{rowMap[index]}": {"NCR": con},
                    }
                else:
                    result[ph] = {
                        **result[ph],
                        f"{rowMap[index]}": {"OPEN": open, "CON": con, "NCR": ncr},
                    }

    for ph in phs:
        phData(ph, phsMap[ph]["rowRange"], phsMap[ph]["rowMap"])
    result["G-TOTAL"] = {
        "NCR": [
            sntzSigV(detailedPuSheet.cell(118, 11).value),
            sntzSigV(detailedPuSheet.cell(118, 12).value),
            sntzSigVPer(detailedPuSheet.cell(118, 13).value),
        ]
    }
    return result


def extract_vital_mod(filePath):
    path = filePath.split("/")[-1]
    wb = load_workbook(filePath, data_only=True)
    vitalSheet = wb[
        f"VITAL (Ear.+Exp.) {path[10:11]}{path[11:13].lower()}-{path[13:15]}"
    ]
    result = {}
    appToDRF = vitalSheet["V41"].value
    appToPF = vitalSheet["V42"].value
    OR = vitalSheet["AJ51"].value
    ORTgtUptoMonth = vitalSheet["AJ50"].value
    ORBud = vitalSheet["AJ49"].value
    result["VITAL"] = {
        "OR": OR,
        "ORBUD": ORBud,
        "ORTGT": ORTgtUptoMonth,
        "APPTODRF": appToDRF,
        "APPTOPF": appToPF,
    }
    return result


def addToDatabase(month):
    registerURL = (
        "https://mydata.apurvasingh.dev/api/v1/telebot/NCRAccountsBot/postData"
    )
    data1 = extractData(f"./files/OWE-{month.upper()}.xlsx")
    payload = {
        "month": f"{month.upper()}",
        "type": "OWE",
        "data1": data1,
    }
    resp = requests.post(registerURL, json=payload)
    return resp.json()


def addSummaryToDatabase(month):
    registerURL = (
        "https://mydata.apurvasingh.dev/api/v1/telebot/NCRAccountsBot/updateData"
    )
    data2 = extractDataSummary(f"./files/OWE-{month.upper()}.xlsx")
    encodedToken = gen_token(token)
    headers = {"token": encodedToken}
    payload = {
        "month": f"{month.upper()}",
        "type": "OWE",
        "data2": data2,
    }
    resp = requests.post(registerURL, json=payload, headers=headers)
    return resp.json()


def addToDatabaseCapex(filePath, sheet):
    registerURL = (
        "https://mydata.apurvasingh.dev/api/v1/telebot/NCRAccountsBot/postData"
    )
    data1 = extractDataCapex(filePath, sheet)
    payload = {
        "month": "JAN22",
        "type": "CAPEX",
        "data1": data1,
    }
    resp = requests.post(registerURL, json=payload)
    return resp.json()


def addToDatabaseCapexUpdate(filePath, sheet):
    registerURL = (
        "https://mydata.apurvasingh.dev/api/v1/telebot/NCRAccountsBot/updateData"
    )
    data1 = extractDataCapex(filePath, sheet)
    encodedToken = gen_token(token)
    headers = {"token": encodedToken}
    payload = {
        "month": "JAN22",
        "type": "CAPEX",
        "data1": data1,
    }
    resp = requests.post(registerURL, json=payload, headers=headers)
    return resp.json()


def updateToDatabase(month):
    registerURL = (
        "https://mydata.apurvasingh.dev/api/v1/telebot/NCRAccountsBot/updateData"
    )
    data1 = extractData(f"./files/OWE-{month.upper()}.xlsx")
    encodedToken = gen_token(token)
    headers = {"token": encodedToken}
    payload = {
        "month": f"{month.upper()}",
        "type": "OWE",
        "data1": data1,
    }
    resp = requests.post(registerURL, json=payload, headers=headers)
    return resp.json()


def updateToDatabaseDiv(month, division):
    registerURL = (
        "https://mydata.apurvasingh.dev/api/v1/telebot/NCRAccountsBot/updateData"
    )
    data3 = extractData(f"./files/OWE-{month.upper()}-{division.upper()}.xlsx")
    encodedToken = gen_token(token)
    headers = {"token": encodedToken}
    payload = {
        "month": f"{month.upper()}",
        "type": "OWE",
        "data3": {f"{division.upper()}": data3},
    }
    resp = requests.post(registerURL, json=payload, headers=headers)
    return resp.json()


def get_owe_data(month):
    dataURL = f"{ncr_data_url}/getData/{month}/OWE"
    encodedToken = gen_token(token)
    headers = {"token": encodedToken}
    res = requests.get(dataURL, headers=headers).json()
    if "monthData" in res.keys():
        return res["monthData"]
    else:
        return res


def postOweMonthlyData(month):
    print(addToDatabase(month))
    print(addSummaryToDatabase(month))
    print(updateToDatabaseDiv(month, "JHS"))
    print(updateToDatabaseDiv(month, "PRYJ"))
    print(updateToDatabaseDiv(month, "AGC"))


if __name__ == "__main__":
    postOweMonthlyData("MAR24")
    print("done")

    postOweMonthlyData("APR24")
    # res = extractData("./files/OWE-Jun22.xlsx")
    # print(res)
    # res = updateToDatabaseDiv("APR22", "PRYJ")
    # print(res)
    print("done")
